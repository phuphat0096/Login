from flask import Flask, request, redirect, url_for
import openpyxl
import os
from datetime import datetime

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, 'dataWeb.xlsx')

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        # Create Registration sheet
        ws_reg = wb.active
        ws_reg.title = "Registration"
        ws_reg.append(['Timestamp', 'Full Name', 'Email', 'Password'])
        
        # Create Usage sheet
        ws_usage = wb.create_sheet("Usage")
        ws_usage.append(['Timestamp', 'Activity', 'Email'])
        
        wb.save(EXCEL_FILE)
    else:
        # Ensure sheets exist if file already exists
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Registration" not in wb.sheetnames:
            ws_reg = wb.create_sheet("Registration")
            ws_reg.append(['Timestamp', 'Full Name', 'Email', 'Password'])
        if "Usage" not in wb.sheetnames:
            ws_usage = wb.create_sheet("Usage")
            ws_usage.append(['Timestamp', 'Activity', 'Email'])
        wb.save(EXCEL_FILE)

@app.route('/success')
def success():
    msg = request.args.get('msg', 'Operation Successful')
    success_path = os.path.join(BASE_DIR, 'success.html')
    with open(success_path, 'r', encoding='utf-8') as f:
        content = f.read()
    return content.replace('{{MESSAGE}}', msg)

@app.route('/login', methods=['POST'])
def login():
    try:
        email = request.form.get('email')
        password = request.form.get('password')
        
        init_excel()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        # Verify credentials
        ws_reg = wb["Registration"]
        user_found = False
        # Iterate through rows starting from row 2 (skipping header)
        for row in ws_reg.iter_rows(min_row=2, values_only=True):
            # Check if email (index 2) and password (index 3) match
            # Row structure: [Timestamp, Full Name, Email, Password]
            if row[2] == email and row[3] == password:
                user_found = True
                break
        
        if not user_found:
            return f"<h1>404 Eror kub aun</h1><p>Invalid email or password. Please <a href='signup.html'>create an account</a>.</p><a href='Webde.html'>Try Again</a>"

        ws = wb["Usage"]
        ws.append([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Login', email])
        wb.save(EXCEL_FILE)
        
        return redirect(url_for('success', msg=f'Welcome back, {email}!'))
    except Exception as e:
        return f"<h1>Error</h1><p>{str(e)}</p><a href='Webde.html'>Go Back</a>"

@app.route('/signup', methods=['POST'])
def signup():
    try:
        fullname = request.form.get('fullname')
        email = request.form.get('email')
        password = request.form.get('password')
        
        init_excel()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Registration"]
        ws.append([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), fullname, email, password])
        wb.save(EXCEL_FILE)
        
        return redirect(url_for('success', msg=f'Account created for {fullname}!'))
    except Exception as e:
        return f"<h1>Error</h1><p>{str(e)}</p><a href='signup.html'>Go Back</a>"

@app.route('/reset-password', methods=['POST'])
def reset_password():
    try:
        email = request.form.get('email')
        new_password = request.form.get('new_password')
        
        init_excel()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_reg = wb["Registration"]
        
        user_found = False
        # Iterate through rows to find the email
        for row in ws_reg.iter_rows(min_row=2):
            # row[2] is email (0-indexed: Timestamp, Full Name, Email, Password)
            if row[2].value == email:
                row[3].value = new_password
                user_found = True
                break
        
        if user_found:
            wb.save(EXCEL_FILE)
            return redirect(url_for('success', msg='Password updated successfully!'))
        else:
            return f"<h1>Error</h1><p>Email not found. Please check your email or <a href='signup.html'>create an account</a>.</p><a href='forgot_password.html'>Try Again</a>"

    except Exception as e:
        return f"<h1>Error</h1><p>{str(e)}</p><a href='forgot_password.html'>Go Back</a>"

# Serve static files so we can open the HTML pages via localhost
from flask import send_from_directory

@app.route('/<path:filename>')
def serve_static(filename):
    return send_from_directory(BASE_DIR, filename)

@app.route('/')
def index():
    return redirect('/Webde.html')

if __name__ == '__main__':
    init_excel()
    print(f"Server running. Saving data to {EXCEL_FILE}")
    app.run(debug=True, port=5000)
